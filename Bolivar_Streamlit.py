"""
=============================================================================
SOLICITUDES DE OUTSOURCING - ARL BOLÍVAR
App Streamlit — versión autónoma en la nube
=============================================================================
Dependencias (requirements.txt):
    streamlit
    supabase
    pandas
    openpyxl
    holidays
    pillow
    reportlab
    smtplib (built-in)

Variables de entorno requeridas (Streamlit Secrets o .env local):
    SUPABASE_URL
    SUPABASE_KEY
    EMAIL_USER
    EMAIL_PASS
    LOGO_URL   ← URL pública de la imagen del logo (subida a Supabase Storage)
=============================================================================
"""

import streamlit as st
import pandas as pd
import holidays
import datetime
import smtplib
import os
import io
import copy
import json
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

import openpyxl
from openpyxl.drawing.image import Image as XLImage

from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import cm, mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT

# from supabase import create_client, Client  # ← DESCOMENTA CUANDO TENGAS SUPABASE LISTO
import urllib.request

# ─────────────────────────────────────────────────────────────────────────────
# 1. CONFIGURACIÓN GENERAL
# ─────────────────────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Solicitudes Outsourcing — ARL Bolívar",
    page_icon="📋",
    layout="wide"
)

# Credenciales desde Streamlit Secrets (en prueba local usa st.secrets o os.environ)
def get_secret(key: str, default: str = "") -> str:
    try:
        return st.secrets[key]
    except Exception:
        return os.getenv(key, default)

SUPABASE_URL  = get_secret("SUPABASE_URL")
SUPABASE_KEY  = get_secret("SUPABASE_KEY")
EMAIL_USER    = get_secret("EMAIL_USER", "notificaciones.bi.adecco@gmail.com")
EMAIL_PASS    = get_secret("EMAIL_PASS")
LOGO_URL      = get_secret("LOGO_URL", "")   # URL pública del logo en Supabase Storage

# ─── SUPABASE DESACTIVADO PARA PRUEBA LOCAL ───────────────────────────────────
# Cuando tengas Supabase listo, descomenta estas líneas y comenta "supabase = None"
# @st.cache_resource
# def init_supabase() -> Client:
#     return create_client(SUPABASE_URL, SUPABASE_KEY)
# supabase: Client = init_supabase()
supabase = None
# ─────────────────────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────────────────────
# 2. LISTAS DE CIUDADES Y GRUPOS
# ─────────────────────────────────────────────────────────────────────────────

GRUPO_A_NEISY = [
    "Apia", "Armenia", "Belen De Umbria", "Circasia", "Dosquebradas", "Manizales",
    "Marmato", "Pereira", "Santa Rosa De Cabal", "Viterbo",
    "Barranquilla", "Cartagena", "Chiriguana", "El Paso", "Galapa", "Malambo",
    "Monteria", "Palmar de Varela", "Puerto Colombia", "Santa Marta", "Santo Tomas",
    "Sincelejo", "Soledad", "Turbaco", "Valledupar",
    "Acacias", "Castilla La Nueva", "Guamal", "Puerto Gaitan", "Villavicencio"
]

GRUPO_B_CAMILA = [
    "Amaga", "Apartado", "Bello", "Buriticá", "Caldas", "Cisneros", "Copacabana",
    "El Carmen de Atrato", "Envigado", "Girardota", "Guarne", "Itagui", "La Estrella",
    "La Union", "Medellin", "Remedios", "Rionegro", "Sabaneta", "San Pedro De Los Milagros",
    "Santa Barbara", "Santafe De Antioquia", "Segovia",
    "Buenaventura", "Cali", "Cartago", "Dagua", "Jamundi", "Palmira", "Yumbo",
    "Popayan"
]

GRUPO_C_JINETH = [
    "Aguazul", "Anapoima", "Barrancabermeja", "Bogota, D.C.", "Bucaramanga", "Cajica",
    "Chia", "Cota", "Cucuta", "Duitama", "Facatativa", "Floridablanca", "Funza",
    "Fusagasuga", "Gachancipa", "Giron", "Granada", "Ibague", "La Calera", "Los Patios",
    "Los Santos", "Madrid", "Mitu", "Mosquera", "Neiva", "Palermo", "Piedecuesta",
    "Quibdo", "Riohacha", "Santa Maria", "Sibate", "Siberia", "Soacha", "Sogamoso",
    "Sopo", "Tenjo", "Tocancipa", "Tunja", "Usaquen", "Villa Rica", "Villanueva",
    "Villapinzon", "Villeta", "Yopal", "Zipaquira"
]

TODAS_LAS_CIUDADES = sorted(GRUPO_A_NEISY + GRUPO_B_CAMILA + GRUPO_C_JINETH)

CORREOS_ASESORAS = {
    "Neisy Bolaños":  "arelis.bolanos@adecco.com",
    "Camila Londoño": "maria.londono@adecco.com",
    "Jineth Cortes":  "jineth.cortes@adecco.com",
}

CC_FIJOS = ["manuel.pimentel@adecco.com", "ingrid.bautista@adecco.com"]

CIUDADES_PRINCIPALES  = ["Bogota, D.C.", "Medellin", "Cali", "Barranquilla",
                          "Cartagena", "Bucaramanga", "Itagui"]
CIUDADES_INTERMEDIAS  = ["Villavicencio", "Neiva", "Ibague", "Pereira",
                          "Manizales", "Armenia", "Cucuta"]

TIEMPOS_RESPUESTA = {"PRINCIPAL": 5, "INTERMEDIA": 7, "ALEJADA": 9}

PROFESIONES = [
    "MÉDICO", "ENFERMERO/A", "FISIOTERAPEUTA", "TERAPEUTA OCUPACIONAL",
    "FONOAUDIÓLOGO/A", "PSICÓLOGO/A", "HIGIENISTA ORAL", "BACTERIÓLOGO/A",
    "NUTRICIONISTA", "INGENIERO AMBIENTAL", "INGENIERO AMBIENTAL Y SANITARIO",
    "INGENIERO DE PROCESOS", "INGENIERO DE PRODUCCIÓN", "INGENIERO ELECTRICISTA",
    "INGENIERO ELECTROMECÁNICO", "INGENIERO INDUSTRIAL", "INGENIERO MECÁNICO",
    "INGENIERO QUÍMICO", "INGENIERO SANITARIO", "OTRO"
]

ESPECIALIDADES = [
    "MÉDICO GENERAL", "MÉDICO ESPECIALISTA SST", "MÉDICO LABORAL",
    "MÉDICO OCUPACIONAL", "OTRO"
]

NIVELES_FORMACION = [
    "TÉCNICO", "TECNÓLOGO", "PROFESIONAL", "ESPECIALISTA", "MAGÍSTER", "OTRO"
]

EXPERIENCIA_ANIOS = [
    "Menos de 2 AÑOS", "De 2 a 5 AÑOS", "De 5 a 9 AÑOS",
    "Mayor a 10 AÑOS", "Otra"
]

DIAS_SEMANA = ["LUNES", "MARTES", "MIÉRCOLES", "JUEVES", "VIERNES", "SÁBADO", "DOMINGO", "FESTIVOS"]

SECTORES_ECONOMICOS = [
    "AGRICULTURA", "COMERCIO", "CONSTRUCCIÓN", "EDUCACIÓN", "FINANCIERO",
    "HIDROCARBUROS", "INDUSTRIA MANUFACTURERA", "MINERÍA", "SALUD",
    "SERVICIOS", "TRANSPORTE", "OTRO"
]

# ─────────────────────────────────────────────────────────────────────────────
# 3. FUNCIONES AUXILIARES
# ─────────────────────────────────────────────────────────────────────────────

def obtener_asesora_y_clasificacion(ciudad: str):
    """Retorna (nombre_asesora, clasificacion, dias_habiles)."""
    if ciudad in GRUPO_A_NEISY:
        asesora = "Neisy Bolaños"
    elif ciudad in GRUPO_B_CAMILA:
        asesora = "Camila Londoño"
    elif ciudad in GRUPO_C_JINETH:
        asesora = "Jineth Cortes"
    else:
        asesora = "Jineth Cortes"  # fallback

    if ciudad in CIUDADES_PRINCIPALES:
        clasificacion = "PRINCIPAL"
    elif ciudad in CIUDADES_INTERMEDIAS:
        clasificacion = "INTERMEDIA"
    else:
        clasificacion = "ALEJADA"

    return asesora, clasificacion, TIEMPOS_RESPUESTA[clasificacion]


def calcular_fecha_entrega(dias_habiles: int) -> str:
    """Calcula la fecha de entrega en días hábiles colombianos desde hoy."""
    hoy = datetime.date.today()
    anios = [hoy.year, hoy.year + 1]
    festivos = holidays.country_holidays("CO", years=anios)

    # Si ya pasó el mediodía, empieza a contar desde mañana
    if datetime.datetime.now().hour >= 12:
        hoy += datetime.timedelta(days=1)
        while hoy.weekday() >= 5 or hoy in festivos:
            hoy += datetime.timedelta(days=1)

    contador = 0
    fecha = hoy
    while contador < dias_habiles:
        if fecha.weekday() < 5 and fecha not in festivos:
            contador += 1
            if contador == dias_habiles:
                break
        fecha += datetime.timedelta(days=1)

    return fecha.strftime("%d/%m/%Y")


def generar_id_solicitud() -> str:
    """Genera un ID único. En producción consulta el último en Supabase."""
    # ── PRODUCCIÓN: descomenta esto cuando Supabase esté activo ──────────────
    # try:
    #     result = supabase.table("solicitudes").select("id_solicitud").order(
    #         "id_solicitud", desc=True).limit(1).execute()
    #     if result.data:
    #         ultimo = int(result.data[0]["id_solicitud"])
    #         return str(ultimo + 1).zfill(4)
    #     return "0001"
    # except Exception:
    #     pass
    # ── PRUEBA LOCAL: ID basado en timestamp ─────────────────────────────────
    return datetime.datetime.now().strftime("%Y%m%d%H%M%S")


# ─────────────────────────────────────────────────────────────────────────────
# 4. GENERACIÓN DEL EXCEL DILIGENCIADO
# ─────────────────────────────────────────────────────────────────────────────

def column_letter_to_index(col: str) -> int:
    index = 0
    for char in col.upper():
        index = index * 26 + (ord(char) - ord("A")) + 1
    return index


def index_to_column_letter(index: int) -> str:
    result = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def get_offset_column(col_letter: str, offset: int) -> str:
    return index_to_column_letter(column_letter_to_index(col_letter) + offset)


def diligenciar_formato_excel(datos: dict, plantilla_bytes: bytes, logo_bytes: bytes | None) -> bytes:
    """
    Toma la plantilla en bytes, la diligencia con los datos del formulario
    y retorna el xlsx resultante en bytes.
    """
    wb = openpyxl.load_workbook(io.BytesIO(plantilla_bytes))
    ws = wb["FORMATO"]

    row = datos  # alias para mantener nombres consistentes con el script original

    def fill_si_no_block(q_map: dict, displacement: int):
        for q, si_cell in q_map.items():
            response = str(row.get(q, "")).strip().upper()
            if response not in ("SI", "NO"):
                continue
            col_letter = "".join(c for c in si_cell if c.isalpha())
            row_number = "".join(c for c in si_cell if c.isdigit())
            no_col_letter = get_offset_column(col_letter, displacement)
            no_cell = no_col_letter + row_number
            if response == "SI":
                ws[si_cell].value = "X"
            elif response == "NO":
                ws[no_cell].value = "X"

    import pandas as pd  # para pd.notna

    # ── Sección I: Información general ──────────────────────────────────────
    ws["F8"].value  = row.get("Q2", "")   # Fecha solicitud (auto)
    if row.get("Q6") == "NUEVO":
        ws["Z8"].value = "X"
    elif row.get("Q6") == "REEMPLAZO":
        ws["AE8"].value = "X"

    ws["P10"].value  = row.get("Q7", "")   # A quién reemplaza
    ws["Q12"].value  = row.get("Q8", "")   # Razón social empresa
    ws["AG12"].value = row.get("Q9", "")   # NIT
    ws["M14"].value  = row.get("Q10", "")  # Nombre AGR
    ws["H16"].value  = row.get("Q11", "")  # Correo AGR
    ws["AF16"].value = row.get("Q12", "")  # Celular AGR
    ws["G18"].value  = row.get("Q13", "")  # Dirección sectorial
    ws["X18"].value  = row.get("Q14", "")  # Nombre DS

    # ── Sección II: Información del cargo ───────────────────────────────────
    formacion = " - ".join(
        str(row[f"Q{i}"]) for i in range(15, 20)
        if row.get(f"Q{i}") and str(row.get(f"Q{i}", "")).strip() not in ("", "nan", "None")
    )
    ws["G23"].value = formacion

    experiencia = " - ".join(
        str(row[f"Q{i}"]) for i in range(20, 22)
        if row.get(f"Q{i}") and str(row.get(f"Q{i}", "")).strip() not in ("", "nan", "None")
    )
    ws["Q29"].value = experiencia

    ws["D35"].value = row.get("Q22", "")   # Salario

    if row.get("Q23") == "FIJO":
        ws["AG35"].value = "X"
    elif row.get("Q23") == "INTERDISCIPLINARIO":
        ws["Z35"].value = "X"

    if row.get("Q35") == "150 HORAS":
        ws["R37"].value = "X"
    elif row.get("Q35") == "75 HORAS":
        ws["L37"].value = "X"

    ws["Z37"].value  = row.get("Q91", "")  # Número de vacantes
    ws["O39"].value  = row.get("Q36", "")  # Ciudad

    # Días de servicio
    dias_raw = str(row.get("Q37", ""))
    dias_seleccionados = [d.strip().upper() for d in dias_raw.split(";") if d.strip()]
    casillas_dias = {
        "LUNES": "H41", "MARTES": "K41", "MIÉRCOLES": "N41", "JUEVES": "Q41",
        "VIERNES": "T41", "SÁBADO": "W41", "DOMINGO": "Z41", "FESTIVOS": "AG41"
    }
    for dia, celda in casillas_dias.items():
        ws[celda].value = "X" if dia in dias_seleccionados else ""

    ws["G43"].value = row.get("Q38", "")  # Horario

    # Clase de riesgo
    opciones_riesgo = {1: "H45", 2: "K45", 3: "M45", 4: "P45", 5: "S45"}
    try:
        riesgo = int(row.get("Q39", 0))
        if riesgo in opciones_riesgo:
            ws[opciones_riesgo[riesgo]].value = "X"
    except Exception:
        pass

    ws["AA45"].value = row.get("Q40", "")  # Sector económico

    # ── Sección 2.1: Distribución interdisciplinaria ─────────────────────────
    ws["H49"].value  = row.get("Q24", "")   # AGR líder
    ws["AC49"].value = row.get("Q25", "")   # Horas AGR líder
    ws["H51"].value  = row.get("Q27", "")   # AGR 1
    ws["AC51"].value = row.get("Q28", "")   # Horas AGR 1
    ws["H53"].value  = row.get("Q30", "")   # AGR 2
    ws["AC53"].value = row.get("Q31", "")   # Horas AGR 2
    ws["H55"].value  = row.get("Q33", "")   # AGR 3
    ws["AC55"].value = row.get("Q34", "")   # Horas AGR 3

    # Transporte propio
    if row.get("Q41") == "MOTO":
        ws["R59"].value = "X"
    elif row.get("Q41") == "VEHÍCULO":
        ws["W59"].value = "X"
    ws["AC59"].value = row.get("Q42", "")   # Auxilio transporte

    # ── Sección 2.2: Auxilios ────────────────────────────────────────────────
    def x(q): return "X" if str(row.get(q, "")).upper() == "SI" else ""
    def no(q): return "X" if str(row.get(q, "")).upper() == "NO" else ""

    ws["I62"].value  = x("Q43");  ws["L62"].value  = no("Q43")
    ws["Q62"].value  = row.get("Q44", "");  ws["AC62"].value = row.get("Q45", "")
    ws["I63"].value  = x("Q46");  ws["L63"].value  = no("Q46")
    ws["Q63"].value  = row.get("Q47", "");  ws["AC63"].value = row.get("Q48", "")
    ws["I65"].value  = x("Q49");  ws["L65"].value  = no("Q49")
    ws["Q65"].value  = row.get("Q50", "");  ws["AC65"].value = row.get("Q51", "")

    ws["H67"].value  = row.get("Q53", "")   # Competencias técnicas
    ws["Q67"].value  = row.get("Q54", "")   # ¿Otra prueba?
    ws["AC67"].value = row.get("Q55", "")   # Descripción prueba

    # ── Sección III: Competencias ────────────────────────────────────────────
    comp_map = {
        "Q56": "O74", "Q57": "O76", "Q58": "O78", "Q59": "O80",
        "Q60": "O82", "Q61": "O84", "Q62": "O86", "Q63": "O88",
        "Q64": "O90", "Q65": "O92",
        "Q66": "AG74", "Q67": "AG76", "Q68": "AG78", "Q69": "AG80"
    }
    fill_si_no_block(comp_map, 3)
    ws["S84"].value = row.get("Q70", "")

    # ── Sección IV: EPPs, dotación, equipos, cursos ──────────────────────────
    epps_map = {
        "Q71": "O98",  "Q72": "O100", "Q73": "O102", "Q74": "O104",
        "Q75": "O106", "Q76": "O108", "Q77": "O110",
        "Q78": "AG98", "Q79": "AG100","Q80": "AG102","Q81": "AG104",
        "Q82": "AG106","Q83": "AG108","Q84": "AG110"
    }
    fill_si_no_block(epps_map, 3)
    ws["U111"].value = row.get("Q85", "")

    extras_map = {
        "Q86": "O114", "Q87": "O116",
        "Q88": "AG114","Q89": "AG116"
    }
    fill_si_no_block(extras_map, 3)

    # ── Sección V: Recomendaciones ───────────────────────────────────────────
    ws["A119"].value = row.get("Q90", "")

    # ── Logo ─────────────────────────────────────────────────────────────────
    if logo_bytes:
        try:
            img = XLImage(io.BytesIO(logo_bytes))
            img.width  = 120
            img.height = 50
            img.anchor = "B2"
            ws.add_image(img)
        except Exception as e:
            st.warning(f"No se pudo insertar el logo en el Excel: {e}")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


# ─────────────────────────────────────────────────────────────────────────────
# 5. GENERACIÓN DEL PDF
# ─────────────────────────────────────────────────────────────────────────────

def generar_pdf(datos: dict, logo_bytes: bytes | None) -> bytes:
    """
    Genera un PDF con el resumen de la solicitud usando ReportLab.
    Replica la estructura visual del FORMATO.xlsx.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm
    )

    styles = getSampleStyleSheet()
    COLOR_VERDE   = colors.HexColor("#1a5276")
    COLOR_VERDE_C = colors.HexColor("#d4e6f1")
    COLOR_GRIS    = colors.HexColor("#f2f3f4")

    estilo_titulo = ParagraphStyle(
        "titulo", parent=styles["Normal"],
        fontSize=13, fontName="Helvetica-Bold",
        textColor=COLOR_VERDE, alignment=TA_CENTER, spaceAfter=4
    )
    estilo_seccion = ParagraphStyle(
        "seccion", parent=styles["Normal"],
        fontSize=9, fontName="Helvetica-Bold",
        textColor=colors.white, alignment=TA_LEFT,
        backColor=COLOR_VERDE, spaceBefore=6, spaceAfter=2,
        leftIndent=4
    )
    estilo_campo = ParagraphStyle(
        "campo", parent=styles["Normal"],
        fontSize=7.5, fontName="Helvetica-Bold", textColor=COLOR_VERDE
    )
    estilo_valor = ParagraphStyle(
        "valor", parent=styles["Normal"],
        fontSize=7.5, fontName="Helvetica"
    )
    estilo_nota = ParagraphStyle(
        "nota", parent=styles["Normal"],
        fontSize=7, fontName="Helvetica-Oblique", textColor=colors.grey
    )

    row = datos
    elementos = []

    # ── Encabezado ────────────────────────────────────────────────────────────
    header_data = [[None, None]]
    if logo_bytes:
        try:
            logo_img = RLImage(io.BytesIO(logo_bytes), width=3*cm, height=1.2*cm)
            header_data = [[logo_img, Paragraph(
                "SOLICITUD DE OUTSOURCING DE SERVICIOS<br/>ESPECIALIZADOS DE GESTIÓN",
                estilo_titulo)]]
        except Exception:
            header_data = [["", Paragraph(
                "SOLICITUD DE OUTSOURCING DE SERVICIOS<br/>ESPECIALIZADOS DE GESTIÓN",
                estilo_titulo)]]
    else:
        header_data = [["", Paragraph(
            "SOLICITUD DE OUTSOURCING DE SERVICIOS<br/>ESPECIALIZADOS DE GESTIÓN",
            estilo_titulo)]]

    tabla_header = Table(header_data, colWidths=[4*cm, 14*cm])
    tabla_header.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("ALIGN",  (1,0), (1,0),  "CENTER"),
        ("BACKGROUND", (0,0), (-1,-1), COLOR_GRIS),
        ("BOX", (0,0), (-1,-1), 0.5, COLOR_VERDE),
        ("TOPPADDING", (0,0), (-1,-1), 6),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
    ]))
    elementos.append(tabla_header)
    elementos.append(Spacer(1, 4))

    # Sub-encabezado: ID, fecha, empresa
    asesora, clasificacion, dias_hab = obtener_asesora_y_clasificacion(str(row.get("Q36", "")))
    dias_totales = dias_hab + 2 if row.get("Q17") == "MÉDICO" else dias_hab
    fecha_entrega = calcular_fecha_entrega(dias_totales)

    sub_data = [
        [Paragraph("ID SOLICITUD", estilo_campo),
         Paragraph(str(row.get("id_solicitud", "")), estilo_valor),
         Paragraph("FECHA", estilo_campo),
         Paragraph(str(row.get("Q2", datetime.date.today().strftime("%d/%m/%Y"))), estilo_valor)],
        [Paragraph("EMPRESA", estilo_campo),
         Paragraph("ARL SEGUROS BOLÍVAR", estilo_valor),
         Paragraph("ASESORA ASIGNADA", estilo_campo),
         Paragraph(asesora, estilo_valor)],
    ]
    tabla_sub = Table(sub_data, colWidths=[3.5*cm, 6*cm, 3.5*cm, 5*cm])
    tabla_sub.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), COLOR_GRIS),
        ("BOX", (0,0), (-1,-1), 0.5, COLOR_VERDE),
        ("INNERGRID", (0,0), (-1,-1), 0.25, colors.lightgrey),
        ("TOPPADDING", (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ("LEFTPADDING", (0,0), (-1,-1), 4),
    ]))
    elementos.append(tabla_sub)
    elementos.append(Spacer(1, 6))

    # ── Helper para tablas de sección ─────────────────────────────────────────
    def seccion(titulo):
        p = Paragraph(f"  {titulo}", estilo_seccion)
        t = Table([[p]], colWidths=[18*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), COLOR_VERDE),
            ("TOPPADDING", (0,0), (-1,-1), 3),
            ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ]))
        return t

    def fila_datos(campo, valor, campo2="", valor2=""):
        return [
            Paragraph(campo, estilo_campo),
            Paragraph(str(valor) if valor else "—", estilo_valor),
            Paragraph(campo2, estilo_campo),
            Paragraph(str(valor2) if valor2 else "", estilo_valor),
        ]

    def tabla_datos(filas, cols=[4.5*cm, 5.5*cm, 4*cm, 4*cm]):
        t = Table(filas, colWidths=cols)
        t.setStyle(TableStyle([
            ("BOX", (0,0), (-1,-1), 0.5, COLOR_VERDE),
            ("INNERGRID", (0,0), (-1,-1), 0.25, colors.lightgrey),
            ("TOPPADDING", (0,0), (-1,-1), 2),
            ("BOTTOMPADDING", (0,0), (-1,-1), 2),
            ("LEFTPADDING", (0,0), (-1,-1), 4),
            ("ROWBACKGROUNDS", (0,0), (-1,-1), [colors.white, COLOR_GRIS]),
        ]))
        return t

    # ── I. Información general ────────────────────────────────────────────────
    elementos.append(seccion("I. INFORMACIÓN GENERAL DE SOLICITUD"))
    filas = [
        fila_datos("Tipo de solicitud", row.get("Q6",""), "¿A quién reemplaza?", row.get("Q7","")),
        fila_datos("Razón social empresa", row.get("Q8",""), "NIT", row.get("Q9","")),
        fila_datos("AGR Solicitante", row.get("Q10",""), "Correo AGR", row.get("Q11","")),
        fila_datos("Celular AGR", row.get("Q12",""), "Dirección Sectorial", row.get("Q13","")),
        fila_datos("Nombre DS", row.get("Q14",""), "", ""),
    ]
    elementos.append(tabla_datos(filas))
    elementos.append(Spacer(1, 4))

    # ── II. Información del cargo ─────────────────────────────────────────────
    elementos.append(seccion("II. INFORMACIÓN GENERAL DEL CARGO"))
    formacion = " - ".join(
        str(row.get(f"Q{i}","")) for i in range(15, 20)
        if str(row.get(f"Q{i}","")).strip() not in ("","nan","None")
    )
    experiencia = " - ".join(
        str(row.get(f"Q{i}","")) for i in range(20, 22)
        if str(row.get(f"Q{i}","")).strip() not in ("","nan","None")
    )
    filas = [
        fila_datos("Formación académica", formacion, "Experiencia requerida", experiencia),
        fila_datos("Salario", row.get("Q22",""), "Tipo asignación", row.get("Q23","")),
        fila_datos("Tiempo de servicio", row.get("Q35",""), "N° vacantes", row.get("Q91","")),
        fila_datos("Ciudad/municipio", row.get("Q36",""), "Horario", row.get("Q38","")),
        fila_datos("Días de servicio", row.get("Q37",""), "Clase de riesgo", row.get("Q39","")),
        fila_datos("Sector económico", row.get("Q40",""), "Transporte propio", row.get("Q41","")),
        fila_datos("Auxilio transporte", row.get("Q42",""), "", ""),
    ]
    elementos.append(tabla_datos(filas))
    elementos.append(Spacer(1, 4))

    # ── 2.1 Distribución interdisciplinaria ───────────────────────────────────
    elementos.append(seccion("2.1. DISTRIBUCIÓN RECURSO INTERDISCIPLINARIO"))
    filas = [
        fila_datos("AGR Líder/Responsable", row.get("Q24",""), "Horas mensuales", row.get("Q25","")),
        fila_datos("AGR 1", row.get("Q27",""), "Horas mensuales", row.get("Q28","")),
        fila_datos("AGR 2", row.get("Q30",""), "Horas mensuales", row.get("Q31","")),
        fila_datos("AGR 3", row.get("Q33",""), "Horas mensuales", row.get("Q34","")),
    ]
    elementos.append(tabla_datos(filas))
    elementos.append(Spacer(1, 4))

    # ── 2.2 Auxilios autorizados ──────────────────────────────────────────────
    elementos.append(seccion("2.2. AUXILIOS AUTORIZADOS"))
    filas = [
        fila_datos("Transporte urbano", row.get("Q43",""), "Frecuencia", row.get("Q44","")),
        fila_datos("Valor transp. urbano", row.get("Q45",""), "Transp. intermunicipal", row.get("Q46","")),
        fila_datos("Frec. intermunicipal", row.get("Q47",""), "Valor intermunicipal", row.get("Q48","")),
        fila_datos("Comunicación", row.get("Q49",""), "Frecuencia comunicación", row.get("Q50","")),
        fila_datos("Valor comunicación", row.get("Q51",""), "Otro auxilio", row.get("Q52","")),
    ]
    elementos.append(tabla_datos(filas))
    elementos.append(Spacer(1, 4))

    # ── III. Competencias ─────────────────────────────────────────────────────
    elementos.append(seccion("III. COMPETENCIAS REQUERIDAS"))
    comp_labels = {
        "Q56": "Orientación al logro", "Q57": "Trabajo en equipo",
        "Q58": "Atención al cliente",  "Q59": "Comunicación efectiva",
        "Q60": "Adaptabilidad",        "Q61": "Pensamiento analítico",
        "Q62": "Innovación",           "Q63": "Manejo del conflicto",
        "Q64": "Neg. y persuasión",    "Q65": "Planificación estratégica",
        "Q66": "Comp. técnica 1",      "Q67": "Comp. técnica 2",
        "Q68": "Comp. técnica 3",      "Q69": "Comp. técnica 4",
    }
    comp_filas = []
    items = list(comp_labels.items())
    for i in range(0, len(items), 2):
        q1, lbl1 = items[i]
        if i+1 < len(items):
            q2, lbl2 = items[i+1]
            comp_filas.append(fila_datos(lbl1, row.get(q1,""), lbl2, row.get(q2,"")))
        else:
            comp_filas.append(fila_datos(lbl1, row.get(q1,""), "", ""))
    comp_filas.append(fila_datos("Prueba técnica específica", row.get("Q70",""), "", ""))
    elementos.append(tabla_datos(comp_filas))
    elementos.append(Spacer(1, 4))

    # ── IV. EPPs / Dotación / Equipos / Cursos ────────────────────────────────
    elementos.append(seccion("IV. DOTACIÓN — EPP's — EQUIPO DE CÓMPUTO — CURSOS"))
    epp_labels = {
        "Q71": "Casco dieléctrico",               "Q72": "Casco con barbuquejo",
        "Q73": "Protector auditivo copa",          "Q74": "Protector auditivo inserción",
        "Q75": "Monogafa antiempañante",           "Q76": "Prot. respiratoria",
        "Q77": "Prot. visual (lente)",             "Q78": "Uniforme anti fluido",
        "Q79": "Chaqueta",                         "Q80": "Camisa",
        "Q81": "Jean",                             "Q82": "Botas dieléct./antidesl.",
        "Q83": "Botas dieléct./antiperforante",    "Q84": "Otro EPP",
    }
    epp_filas = []
    items_epp = list(epp_labels.items())
    for i in range(0, len(items_epp), 2):
        q1, lbl1 = items_epp[i]
        if i+1 < len(items_epp):
            q2, lbl2 = items_epp[i+1]
            epp_filas.append(fila_datos(lbl1, row.get(q1,""), lbl2, row.get(q2,"")))
        else:
            epp_filas.append(fila_datos(lbl1, row.get(q1,""), "", ""))
    epp_filas.append(fila_datos("¿Otro EPP? ¿Cuál?", row.get("Q85",""), "", ""))
    epp_filas.append(fila_datos("Equipo cómputo básico", row.get("Q86",""),
                                "Equipo cómputo mayor cap.", row.get("Q87","")))
    epp_filas.append(fila_datos("Curso trabajo en alturas", row.get("Q88",""),
                                "Curso espacios confinados", row.get("Q89","")))
    elementos.append(tabla_datos(epp_filas))
    elementos.append(Spacer(1, 4))

    # ── V. Recomendaciones ────────────────────────────────────────────────────
    elementos.append(seccion("V. RECOMENDACIONES PARA TENER EN CUENTA DURANTE EL PROCESO"))
    rec_data = [[Paragraph(str(row.get("Q90","—")), estilo_valor)]]
    tabla_rec = Table(rec_data, colWidths=[18*cm])
    tabla_rec.setStyle(TableStyle([
        ("BOX", (0,0), (-1,-1), 0.5, COLOR_VERDE),
        ("TOPPADDING", (0,0), (-1,-1), 6),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
        ("LEFTPADDING", (0,0), (-1,-1), 6),
        ("BACKGROUND", (0,0), (-1,-1), COLOR_GRIS),
    ]))
    elementos.append(tabla_rec)
    elementos.append(Spacer(1, 8))

    # ── Pie de página con info de tiempos ─────────────────────────────────────
    pie_data = [[
        Paragraph(f"Clasificación ciudad: <b>{clasificacion}</b>  |  "
                  f"Tiempo de respuesta: <b>{dias_totales} días hábiles</b>  |  "
                  f"Fecha tentativa de entrega: <b>{fecha_entrega}</b>",
                  estilo_nota)
    ]]
    tabla_pie = Table(pie_data, colWidths=[18*cm])
    tabla_pie.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), COLOR_VERDE_C),
        ("BOX", (0,0), (-1,-1), 0.5, COLOR_VERDE),
        ("TOPPADDING", (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ("LEFTPADDING", (0,0), (-1,-1), 6),
    ]))
    elementos.append(tabla_pie)

    doc.build(elementos)
    buffer.seek(0)
    return buffer.read()


# ─────────────────────────────────────────────────────────────────────────────
# 6. ENVÍO DE CORREO
# ─────────────────────────────────────────────────────────────────────────────

def enviar_correo(datos: dict, pdf_bytes: bytes, xlsx_bytes: bytes) -> bool:
    """Envía el correo a la asesora correspondiente con el PDF adjunto."""

    ciudad       = str(datos.get("Q36", "")).strip()
    asesora, clasificacion, dias_hab = obtener_asesora_y_clasificacion(ciudad)
    dias_totales = dias_hab + 2 if datos.get("Q17") == "MÉDICO" else dias_hab
    fecha_entrega = calcular_fecha_entrega(dias_totales)

    correo_agr   = datos.get("Q11", "")
    id_sol       = datos.get("id_solicitud", "N/A")

    formacion = " - ".join(
        str(datos.get(f"Q{i}","")) for i in range(15, 20)
        if str(datos.get(f"Q{i}","")).strip() not in ("","nan","None")
    )

    cuerpo = f"""
    <html><body>
    <p>Estimada Asesora,</p>
    <p>Se adjunta el formulario de Solicitud de Outsourcing para el perfil:</p>
    <p><b>{formacion}</b> en <b>{ciudad}</b>.</p>

    <p><b>NOTA IMPORTANTE:</b> Según la clasificación de la ciudad (<b>{clasificacion}</b>),
    el tiempo de respuesta para esta solicitud es de <b>{dias_totales} días hábiles</b>.</p>
    <p>Por lo tanto la fecha tentativa de entrega será el <b>{fecha_entrega}</b>.</p>
    <p>Número de vacantes solicitadas: <b>{datos.get("Q91","")}</b>.</p>

    <p>Por favor, revisarlo y dar continuidad al proceso.</p>
    <br>
    <p><b>Información del Solicitante:</b></p>
    <ul>
      <li><b>AGR:</b> {datos.get("Q10","")}</li>
      <li><b>Correo AGR:</b> {correo_agr}</li>
      <li><b>Tipo de Afiliación:</b> {datos.get("Q23","")}</li>
      <li><b>Empresa Afiliada:</b> {datos.get("Q8","")}</li>
    </ul>
    <br>
    <p>Atentamente,<br>Equipo de Notificaciones BI.</p>
    </body></html>
    """

    destinatario = CORREOS_ASESORAS.get(asesora, "jineth.cortes@adecco.com")
    cc_lista     = CC_FIJOS + ([correo_agr] if correo_agr else [])
    nombre_base  = f"{id_sol} - {formacion} en {ciudad}"

    msg = MIMEMultipart("alternative")
    msg["From"]    = EMAIL_USER
    msg["To"]      = destinatario
    msg["Cc"]      = "; ".join(cc_lista)
    msg["Subject"] = f"ID {id_sol} - NUEVA SOLICITUD DE OUTSOURCING: {formacion} en {ciudad}"
    msg.attach(MIMEText(cuerpo, "html"))

    # Adjuntar PDF
    parte_pdf = MIMEBase("application", "octet-stream")
    parte_pdf.set_payload(pdf_bytes)
    encoders.encode_base64(parte_pdf)
    parte_pdf.add_header("Content-Disposition", "attachment",
                         filename=("utf-8", "", f"{nombre_base}.pdf"))
    msg.attach(parte_pdf)

    # Adjuntar Excel (opcional — se guarda también en Supabase Storage)
    parte_xlsx = MIMEBase("application", "octet-stream")
    parte_xlsx.set_payload(xlsx_bytes)
    encoders.encode_base64(parte_xlsx)
    parte_xlsx.add_header("Content-Disposition", "attachment",
                          filename=("utf-8", "", f"{nombre_base}.xlsx"))
    msg.attach(parte_xlsx)

    try:
        todos_destinatarios = [destinatario] + cc_lista
        with smtplib.SMTP("smtp.gmail.com", 587) as servidor:
            servidor.starttls()
            servidor.login(EMAIL_USER, EMAIL_PASS)
            servidor.sendmail(EMAIL_USER, todos_destinatarios, msg.as_string())
        return True
    except Exception as e:
        st.error(f"Error al enviar correo: {e}")
        return False


# ─────────────────────────────────────────────────────────────────────────────
# 7. SUPABASE: GUARDAR DATOS Y ARCHIVOS
# ─────────────────────────────────────────────────────────────────────────────

def guardar_solicitud_supabase(datos: dict) -> bool:
    """Guarda los datos del formulario en la tabla 'solicitudes'.
    DESACTIVADO EN PRUEBA LOCAL — descomenta el cuerpo cuando Supabase esté listo."""
    # try:
    #     registro = {k: (None if str(v).strip() in ("nan","None","") else v)
    #                 for k, v in datos.items()}
    #     supabase.table("solicitudes").insert(registro).execute()
    #     return True
    # except Exception as e:
    #     st.error(f"Error guardando en Supabase (solicitudes): {e}")
    #     return False
    return True


def registrar_envio_supabase(datos: dict, asesora: str, exito: bool) -> bool:
    """Guarda el registro del envío en la tabla 'envios'.
    DESACTIVADO EN PRUEBA LOCAL — descomenta el cuerpo cuando Supabase esté listo."""
    # try:
    #     registro = {
    #         "id_solicitud":  datos.get("id_solicitud"),
    #         "asesora":       asesora,
    #         "correo_asesora": CORREOS_ASESORAS.get(asesora,""),
    #         "ciudad":        datos.get("Q36",""),
    #         "perfil":        datos.get("Q15",""),
    #         "fecha_envio":   datetime.datetime.now().isoformat(),
    #         "estado":        "ENVIADO" if exito else "ERROR",
    #     }
    #     supabase.table("envios").insert(registro).execute()
    #     return True
    # except Exception as e:
    #     st.error(f"Error guardando en Supabase (envios): {e}")
    #     return False
    return True


def subir_archivo_supabase(file_bytes: bytes, path: str, content_type: str) -> bool:
    """Sube un archivo al bucket 'solicitudes-bolivar' en Supabase Storage.
    DESACTIVADO EN PRUEBA LOCAL — descomenta el cuerpo cuando Supabase esté listo."""
    # try:
    #     supabase.storage.from_("solicitudes-bolivar").upload(
    #         path=path,
    #         file=file_bytes,
    #         file_options={"content-type": content_type}
    #     )
    #     return True
    # except Exception as e:
    #     st.warning(f"No se pudo subir archivo a Storage: {e}")
    #     return False
    return True


def obtener_logo() -> bytes | None:
    """Descarga el logo desde la URL pública configurada."""
    if not LOGO_URL:
        return None
    try:
        with urllib.request.urlopen(LOGO_URL) as resp:
            return resp.read()
    except Exception:
        return None


# ─────────────────────────────────────────────────────────────────────────────
# 8. INTERFAZ STREAMLIT
# ─────────────────────────────────────────────────────────────────────────────

def main():
    # ── CSS personalizado ─────────────────────────────────────────────────────
    st.markdown("""
    <style>
    .stApp { background-color: #f8f9fa; }
    div[data-testid="stForm"] {
        background: white;
        padding: 24px;
        border-radius: 12px;
        border: 1px solid #dee2e6;
        box-shadow: 0 2px 8px rgba(0,0,0,0.06);
    }
    .section-header {
        background: #1a5276;
        color: white;
        padding: 8px 14px;
        border-radius: 6px;
        font-weight: 600;
        font-size: 0.95rem;
        margin: 20px 0 10px 0;
    }
    .info-box {
        background: #d4e6f1;
        border-left: 4px solid #1a5276;
        padding: 10px 14px;
        border-radius: 4px;
        font-size: 0.88rem;
        margin: 8px 0;
    }
    </style>
    """, unsafe_allow_html=True)

    # ── Título ────────────────────────────────────────────────────────────────
    col_logo, col_titulo = st.columns([1, 4])
    with col_titulo:
        st.markdown("## 📋 Solicitud de Outsourcing de Servicios Especializados de Gestión")
        st.markdown("**ARL Seguros Bolívar** — Complete el formulario y presione **Enviar Solicitud**.")

    st.divider()

    # ── Formulario ────────────────────────────────────────────────────────────
    with st.form("solicitud_form", clear_on_submit=False):

        # ── I. Información general ────────────────────────────────────────────
        st.markdown('<div class="section-header">I. INFORMACIÓN GENERAL DE SOLICITUD</div>',
                    unsafe_allow_html=True)

        c1, c2 = st.columns(2)
        with c1:
            Q6  = st.selectbox("Tipo de solicitud *", ["NUEVO", "REEMPLAZO"])
        with c2:
            Q7  = st.text_input("En caso de reemplazo, ¿a quién reemplaza?")

        c1, c2 = st.columns(2)
        with c1:
            Q8  = st.text_input("Razón social empresa afiliada *")
        with c2:
            Q9  = st.text_input("NIT de la empresa *")

        c1, c2, c3 = st.columns(3)
        with c1:
            Q10 = st.text_input("AGR Solicitante (Líder o Responsable) *")
        with c2:
            Q11 = st.text_input("Correo electrónico AGR *")
        with c3:
            Q12 = st.text_input("Número de celular AGR *")

        c1, c2 = st.columns(2)
        with c1:
            Q13 = st.text_input("Dirección Sectorial *")
        with c2:
            Q14 = st.text_input("Nombre DS *")

        # ── II. Información del cargo ─────────────────────────────────────────
        st.markdown('<div class="section-header">II. INFORMACIÓN GENERAL DEL CARGO</div>',
                    unsafe_allow_html=True)

        c1, c2 = st.columns(2)
        with c1:
            Q15 = st.selectbox("Profesión / Perfil principal *", PROFESIONES)
        with c2:
            Q16 = st.text_input("Especialidad médica (si aplica)")

        c1, c2 = st.columns(2)
        with c1:
            Q17 = st.selectbox("Nivel de formación *", NIVELES_FORMACION)
        with c2:
            Q18 = st.text_input("Título específico requerido")

        Q19 = st.text_input("Otra formación / certificación adicional")

        c1, c2 = st.columns(2)
        with c1:
            Q20 = st.selectbox("Experiencia requerida (años mínimos) *", EXPERIENCIA_ANIOS)
        with c2:
            Q21 = st.text_input("Mencione experiencia diferente a lo anterior")

        c1, c2, c3 = st.columns(3)
        with c1:
            Q22 = st.number_input("Salario *", min_value=0, step=50000, value=0)
        with c2:
            Q23 = st.selectbox("Asignación *", ["INTERDISCIPLINARIO", "FIJO"])
        with c3:
            Q91 = st.number_input("Número de vacantes *", min_value=1, step=1, value=1)

        c1, c2 = st.columns(2)
        with c1:
            Q35 = st.selectbox("Tiempo de prestación del servicio *", ["75 HORAS", "150 HORAS"])
        with c2:
            Q36 = st.selectbox("Ciudad/Municipio donde se prestará el servicio *",
                               TODAS_LAS_CIUDADES)

        c1, c2 = st.columns(2)
        with c1:
            Q37_dias = st.multiselect("Días de servicio *", DIAS_SEMANA)
        with c2:
            Q38 = st.text_input("Horario de servicio *")

        c1, c2 = st.columns(2)
        with c1:
            Q39 = st.selectbox("Clase de riesgo *", [1, 2, 3, 4, 5])
        with c2:
            Q40 = st.selectbox("Sector económico *", SECTORES_ECONOMICOS)

        c1, c2 = st.columns(2)
        with c1:
            Q41 = st.selectbox("¿Requiere transporte propio?", ["NINGUNO", "MOTO", "VEHÍCULO"])
        with c2:
            Q42 = st.number_input("Auxilio transporte propio ($)", min_value=0, step=10000, value=0)

        # ── 2.1 Distribución interdisciplinaria ───────────────────────────────
        st.markdown('<div class="section-header">2.1. DISTRIBUCIÓN RECURSO INTERDISCIPLINARIO</div>',
                    unsafe_allow_html=True)

        c1, c2 = st.columns(2)
        with c1:
            Q24 = st.text_input("AGR Líder o Responsable")
        with c2:
            Q25 = st.number_input("Horas mensuales (Líder)", min_value=0, step=1, value=0)

        agregar_agr2 = st.checkbox("Adicionar AGR 1")
        Q27, Q28 = "", 0
        if agregar_agr2:
            c1, c2 = st.columns(2)
            with c1: Q27 = st.text_input("AGR 1")
            with c2: Q28 = st.number_input("Horas mensuales (AGR 1)", min_value=0, step=1, value=0)

        agregar_agr3 = st.checkbox("Adicionar AGR 2")
        Q30, Q31 = "", 0
        if agregar_agr3:
            c1, c2 = st.columns(2)
            with c1: Q30 = st.text_input("AGR 2")
            with c2: Q31 = st.number_input("Horas mensuales (AGR 2)", min_value=0, step=1, value=0)

        agregar_agr4 = st.checkbox("Adicionar AGR 3")
        Q33, Q34 = "", 0
        if agregar_agr4:
            c1, c2 = st.columns(2)
            with c1: Q33 = st.text_input("AGR 3")
            with c2: Q34 = st.number_input("Horas mensuales (AGR 3)", min_value=0, step=1, value=0)

        # ── 2.2 Auxilios autorizados ──────────────────────────────────────────
        st.markdown('<div class="section-header">2.2. AUXILIOS AUTORIZADOS</div>',
                    unsafe_allow_html=True)

        c1, c2, c3 = st.columns(3)
        with c1: Q43 = st.selectbox("Transporte Urbano", ["NO", "SI"])
        with c2: Q44 = st.selectbox("Frecuencia T. Urbano", ["", "QUINCENAL","MENSUAL","ANUAL"])
        with c3: Q45 = st.number_input("Valor T. Urbano ($)", min_value=0, step=10000, value=0)

        c1, c2, c3 = st.columns(3)
        with c1: Q46 = st.selectbox("Transporte Intermunicipal", ["NO", "SI"])
        with c2: Q47 = st.selectbox("Frecuencia T. Intermunicipal", ["", "QUINCENAL","MENSUAL","ANUAL"])
        with c3: Q48 = st.number_input("Valor T. Intermunicipal ($)", min_value=0, step=10000, value=0)

        c1, c2, c3 = st.columns(3)
        with c1: Q49 = st.selectbox("Comunicación", ["NO", "SI"])
        with c2: Q50 = st.selectbox("Frecuencia Comunicación", ["", "QUINCENAL","MENSUAL","ANUAL"])
        with c3: Q51 = st.number_input("Valor Comunicación ($)", min_value=0, step=10000, value=0)

        c1, c2, c3 = st.columns(3)
        with c1: Q52 = st.selectbox("Otro auxilio", ["NO", "SI"])
        with c2: Q53_texto = st.text_input("¿Cuál otro auxilio?")
        with c3: Q54_frec = st.selectbox("Frecuencia otro auxilio", ["", "QUINCENAL","MENSUAL","ANUAL"])
        Q55_valor = st.number_input("Valor otro auxilio ($)", min_value=0, step=10000, value=0)

        # ── III. Competencias técnicas ────────────────────────────────────────
        st.markdown('<div class="section-header">III. COMPETENCIAS TÉCNICAS REQUERIDAS</div>',
                    unsafe_allow_html=True)

        c1, c2 = st.columns(2)
        with c1:
            Q56 = st.selectbox("Prueba técnica específica según profesión", ["NO","SI"])
            Q57 = st.selectbox("Prueba ofimática (mínimo Excel intermedio)", ["NO","SI"])
            Q58 = st.selectbox("Prueba técnica SIG", ["NO","SI"])
        with c2:
            Q53 = st.text_area("Competencias técnicas generales (describe brevemente)")
            Q54 = st.selectbox("¿Otra prueba técnica?", ["NO","SI"])
            Q55 = st.text_input("Describe la prueba solicitada")

        # ── Competencias comportamentales ─────────────────────────────────────
        st.markdown("**Competencias comportamentales requeridas:**")
        comp_labels_form = [
            ("Q59","Orientación al logro"), ("Q60","Trabajo en equipo"),
            ("Q61","Atención al cliente"),  ("Q62","Comunicación efectiva"),
            ("Q63","Adaptabilidad"),        ("Q64","Pensamiento analítico"),
            ("Q65","Innovación"),           ("Q66","Manejo del conflicto"),
            ("Q67","Negociación y persuasión"), ("Q68","Desarrollo de relaciones"),
            ("Q69","Liderar equipos"),      ("Q70","Planificación estratégica"),
        ]
        comp_vals = {}
        for i in range(0, len(comp_labels_form), 3):
            cols = st.columns(3)
            for j, (qkey, qlabel) in enumerate(comp_labels_form[i:i+3]):
                with cols[j]:
                    comp_vals[qkey] = st.selectbox(qlabel, ["NO","SI"], key=f"comp_{qkey}")

        Q70_extra = st.text_input("Descripción adicional competencias técnicas específicas")

        # ── IV. EPPs / Dotación / Equipos / Cursos ────────────────────────────
        st.markdown('<div class="section-header">IV. DOTACIÓN — EPP\'s — EQUIPO DE CÓMPUTO — CURSOS ADICIONALES</div>',
                    unsafe_allow_html=True)

        st.markdown("**4.1. EPP's requeridos:**")
        epp_items = [
            ("Q71","Casco dieléctrico"), ("Q72","Casco con barbuquejo"),
            ("Q73","Protector auditivo de copa"), ("Q74","Protector auditivo de inserción"),
            ("Q75","Monogafa antiempañante"), ("Q76","Protección respiratoria"),
            ("Q77","Protección visual (lente claro/oscuro)"),
        ]
        epp_vals = {}
        for i in range(0, len(epp_items), 3):
            cols = st.columns(3)
            for j, (qk, ql) in enumerate(epp_items[i:i+3]):
                with cols[j]:
                    epp_vals[qk] = st.selectbox(ql, ["NO","SI"], key=f"epp_{qk}")

        st.markdown("**4.2. Dotación:**")
        dot_items = [
            ("Q78","Uniforme anti fluido"), ("Q79","Chaqueta"),
            ("Q80","Camisa"), ("Q81","Jean"),
            ("Q82","Botas dieléct./antideslizante"), ("Q83","Botas dieléct./antiperforante"),
            ("Q84","Otro elemento dotación"),
        ]
        dot_vals = {}
        for i in range(0, len(dot_items), 3):
            cols = st.columns(3)
            for j, (qk, ql) in enumerate(dot_items[i:i+3]):
                with cols[j]:
                    dot_vals[qk] = st.selectbox(ql, ["NO","SI"], key=f"dot_{qk}")

        Q85 = st.text_input("¿Cuál otro elemento de dotación/EPP?")

        st.markdown("**4.3. Equipo de cómputo:**")
        c1, c2 = st.columns(2)
        with c1: Q86 = st.selectbox("Equipo de cómputo básico", ["NO","SI"])
        with c2: Q87 = st.selectbox("Equipo de cómputo mayor capacidad", ["NO","SI"])

        st.markdown("**4.4. Cursos especiales:**")
        c1, c2 = st.columns(2)
        with c1: Q88 = st.selectbox("Curso trabajo seguro en alturas", ["NO","SI"])
        with c2: Q89 = st.selectbox("Curso espacios confinados", ["NO","SI"])

        # ── V. Recomendaciones ────────────────────────────────────────────────
        st.markdown('<div class="section-header">V. RECOMENDACIONES PARA TENER EN CUENTA DURANTE EL PROCESO</div>',
                    unsafe_allow_html=True)
        Q90 = st.text_area("Recomendaciones generales")

        # ── Botón de envío ────────────────────────────────────────────────────
        st.divider()
        submitted = st.form_submit_button(
            "📨 Enviar Solicitud",
            use_container_width=True,
            type="primary"
        )

    # ─────────────────────────────────────────────────────────────────────────
    # 9. PROCESAMIENTO AL ENVIAR
    # ─────────────────────────────────────────────────────────────────────────
    if submitted:
        # Validaciones básicas
        errores = []
        if not Q8.strip():  errores.append("Razón social empresa es obligatorio.")
        if not Q10.strip(): errores.append("Nombre del AGR es obligatorio.")
        if not Q11.strip(): errores.append("Correo del AGR es obligatorio.")
        if not Q36:         errores.append("Debe seleccionar una ciudad.")
        if not Q37_dias:    errores.append("Debe seleccionar al menos un día de servicio.")

        if errores:
            for e in errores:
                st.error(e)
            st.stop()

        with st.spinner("Procesando solicitud..."):

            # Generar ID
            id_sol = generar_id_solicitud()
            fecha_hoy = datetime.date.today().strftime("%d/%m/%Y")

            # Consolidar datos
            datos = {
                "id_solicitud": id_sol,
                "Q2":  fecha_hoy,
                "Q6":  Q6,  "Q7":  Q7,  "Q8":  Q8,  "Q9":  Q9,
                "Q10": Q10, "Q11": Q11, "Q12": Q12, "Q13": Q13, "Q14": Q14,
                "Q15": Q15, "Q16": Q16, "Q17": Q17, "Q18": Q18, "Q19": Q19,
                "Q20": Q20, "Q21": Q21, "Q22": Q22, "Q23": Q23,
                "Q24": Q24, "Q25": Q25,
                "Q27": Q27, "Q28": Q28,
                "Q30": Q30, "Q31": Q31,
                "Q33": Q33, "Q34": Q34,
                "Q35": Q35, "Q36": Q36,
                "Q37": "; ".join(Q37_dias),
                "Q38": Q38, "Q39": Q39, "Q40": Q40, "Q41": Q41, "Q42": Q42,
                "Q43": Q43, "Q44": Q44, "Q45": Q45,
                "Q46": Q46, "Q47": Q47, "Q48": Q48,
                "Q49": Q49, "Q50": Q50, "Q51": Q51,
                "Q52": Q52, "Q53": Q53_texto, "Q54": Q54_frec, "Q55": str(Q55_valor),
                "Q56": Q56, "Q57": Q57, "Q58": Q58,
                **comp_vals,
                "Q70": Q70_extra,
                **epp_vals,
                **dot_vals,
                "Q85": Q85,
                "Q86": Q86, "Q87": Q87,
                "Q88": Q88, "Q89": Q89,
                "Q90": Q90,
                "Q91": Q91,
            }

            # Obtener logo
            logo_bytes = obtener_logo()

            # Leer plantilla Excel del repositorio
            plantilla_path = os.path.join(os.path.dirname(__file__), "FORMATO.xlsx")
            try:
                with open(plantilla_path, "rb") as f:
                    plantilla_bytes = f.read()
            except FileNotFoundError:
                st.error("No se encontró FORMATO.xlsx junto al script. Asegúrate de incluirlo en el repositorio.")
                st.stop()

            # Generar archivos
            xlsx_bytes = diligenciar_formato_excel(datos, plantilla_bytes, logo_bytes)
            pdf_bytes  = generar_pdf(datos, logo_bytes)

            # Guardar en Supabase
            guardar_solicitud_supabase(datos)

            asesora, _, _ = obtener_asesora_y_clasificacion(Q36)
            nombre_base   = f"{id_sol} - {Q15} en {Q36}"
            carpeta       = asesora

            subir_archivo_supabase(
                xlsx_bytes,
                f"{carpeta}/{nombre_base}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            subir_archivo_supabase(
                pdf_bytes,
                f"{carpeta}/{nombre_base}.pdf",
                "application/pdf"
            )

            # Enviar correo
            exito_correo = enviar_correo(datos, pdf_bytes, xlsx_bytes)
            registrar_envio_supabase(datos, asesora, exito_correo)

        # ── Feedback al usuario ───────────────────────────────────────────────
        if exito_correo:
            st.success(f"✅ Solicitud **{id_sol}** enviada exitosamente a **{asesora}**.")
        else:
            st.warning(f"⚠️ Solicitud registrada como **{id_sol}** pero hubo un problema al enviar el correo. Contacta al administrador.")

        # Información calculada
        asesora_info, clasificacion_info, dias_info = obtener_asesora_y_clasificacion(Q36)
        dias_totales = dias_info + 2 if Q17 == "MÉDICO" else dias_info
        fecha_entrega = calcular_fecha_entrega(dias_totales)

        st.markdown(f"""
        <div class="info-box">
        📍 <b>Ciudad:</b> {Q36} ({clasificacion_info}) &nbsp;|&nbsp;
        ⏱️ <b>Tiempo de respuesta:</b> {dias_totales} días hábiles &nbsp;|&nbsp;
        📅 <b>Fecha tentativa de entrega:</b> {fecha_entrega}
        </div>
        """, unsafe_allow_html=True)

        # Botones de descarga
        col1, col2 = st.columns(2)
        with col1:
            st.download_button(
                "⬇️ Descargar PDF",
                data=pdf_bytes,
                file_name=f"{nombre_base}.pdf",
                mime="application/pdf"
            )
        with col2:
            st.download_button(
                "⬇️ Descargar Excel",
                data=xlsx_bytes,
                file_name=f"{nombre_base}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )


if __name__ == "__main__":
    main()